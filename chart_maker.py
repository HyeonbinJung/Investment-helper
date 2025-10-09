from datetime import date, timedelta
import xlwings as xw
import yfinance as yf
import time
from openpyxl.utils import get_column_letter, column_index_from_string
import pandas as pd

wb = xw.Book(r"C:\Users\sjung\OneDrive\바탕 화면\새 폴더 (4)\trading_helper.xlsx")

def days_loader(days, page, cell_Alpha, cell_num):
    ws = wb.sheets[page]
    col_num = column_index_from_string(cell_Alpha)
    title = str(cell_Alpha)+str(cell_num)
    ws[title].value = "Date"
    ws[title].api.Font.Bold = True
    ws[title].api.Font.Italic = True
    for i in range(days):
        reverse = days - i 
        alpha = get_column_letter(col_num + 1 + i)
        insert_cell = alpha + str(cell_num)
        ws[insert_cell].value = (date.today() - timedelta(days=reverse - 1))
        ws[insert_cell].api.HorizontalAlignment = -4108
        ws[insert_cell].api.Font.Bold = True
        ws[insert_cell].api.Font.Italic = True

# Price type: Open(시작가) / High(고점) / Low(저점) / Close(종가) / Volume(거래량)
# example: get_chart(30, 4, "A", 2, "TSLA", "Close")
def get_chart(days, page, cell_Alpha, cell_num, ticker, price_type):
    ws = wb.sheets[page]
    col_num = column_index_from_string(cell_Alpha)
    stock = yf.Ticker(ticker)
    ws[str(cell_Alpha)+str(cell_num)].value = str(price_type)+ " (" +str(ticker) + ")"
    ws[str(cell_Alpha)+str(cell_num)].api.Font.Bold = True
    ws[str(cell_Alpha)+str(cell_num)].api.Font.Italic = True
    for i in range(days):
        reverse = days - i 
        alpha = get_column_letter(col_num + 1 + i)
        stock_cell = alpha + str(cell_num)
        target_date = (date.today() - timedelta(days=reverse - 1))
        target_date_2 = (date.today() - timedelta(days=reverse - 2))
        try:
            stock_price = stock.history(start=target_date, end=target_date_2)[price_type].iloc[0]
            ws[stock_cell].value = stock.history(start=target_date, end=target_date_2)[price_type].iloc[0]
            ws[stock_cell].api.Borders.Weight = 2
        except Exception as e:
            ws[stock_cell].value = False
            ws[stock_cell].api.Borders.Weight = 2


# High price - low price to check fluctuation
def get_chart_advance_fluctuation_range1(days, page, cell_Alpha, cell_num, ticker):
    ws = wb.sheets[page]
    col_num = column_index_from_string(cell_Alpha)
    stock = yf.Ticker(ticker)
    ws[str(cell_Alpha)+str(cell_num)].value = "Fluctuation(High - Low)"
    ws[str(cell_Alpha)+str(cell_num)].api.Font.Bold = True
    ws[str(cell_Alpha)+str(cell_num)].api.Font.Italic = True
    for i in range(days):
        reverse = days - i 
        alpha = get_column_letter(col_num + 1 + i)
        stock_cell = alpha + str(cell_num)
        target_date = (date.today() - timedelta(days=reverse - 1))
        target_date_2 = (date.today() - timedelta(days=reverse - 2))
        try:
            high = stock.history(start=target_date, end=target_date_2)["High"].iloc[0]
            low = stock.history(start=target_date, end=target_date_2)["Low"].iloc[0]
            ws[stock_cell].value = high - low
            ws[stock_cell].api.Borders.Weight = 2
        except Exception as e:
            ws[stock_cell].value = False
            ws[stock_cell].api.Borders.Weight = 2

# High price / low price to check fluctuation
def get_chart_advance_fluctuation_range2(days, page, cell_Alpha, cell_num, ticker):
    ws = wb.sheets[page]
    col_num = column_index_from_string(cell_Alpha)
    stock = yf.Ticker(ticker)
    ws[str(cell_Alpha)+str(cell_num)].value = "Fluctuation(High / Low)"
    ws[str(cell_Alpha)+str(cell_num)].api.Font.Bold = True
    ws[str(cell_Alpha)+str(cell_num)].api.Font.Italic = True
    fluct = 0.00
    no_price_day = 0
    for i in range(days):
        reverse = days - i 
        alpha = get_column_letter(col_num + 1 + i)
        stock_cell = alpha + str(cell_num)
        target_date = (date.today() - timedelta(days=reverse - 1))
        target_date_2 = (date.today() - timedelta(days=reverse - 2))
        try:
            high = stock.history(start=target_date, end=target_date_2)["High"].iloc[0]
            low = stock.history(start=target_date, end=target_date_2)["Low"].iloc[0]
            ws[stock_cell].value = high / low
            fluct = fluct + high / low
            if high / low > 1.30:
                ws[stock_cell].api.Font.Bold = True
            elif high / low > 1.10:
                ws[stock_cell].api.Font.Bold = True
            elif high / low > 1.05:
                ws[stock_cell].api.Font.Bold = True
            ws[stock_cell].api.Borders.Weight = 2
        except Exception as e:
            ws[stock_cell].value = False
            no_price_day = no_price_day + 1
            ws[stock_cell].api.Borders.Weight = 2
    print("Average Fluctuation(High Price/Low Price), given period: "+str(fluct/(days - no_price_day)))

def get_all(days, page, cell_Alpha, cell_num, ticker):
    days_loader(days, page, cell_Alpha, cell_num)
    get_chart(days, page, cell_Alpha, cell_num+2, ticker, "Open")
    get_chart(days, page, cell_Alpha, cell_num+3, ticker, "Close")
    get_chart(days, page, cell_Alpha, cell_num+5, ticker, "High")
    get_chart(days, page, cell_Alpha, cell_num+6, ticker, "Low")
    get_chart(days, page, cell_Alpha, cell_num+8, ticker, "Volume")

    get_chart_advance_fluctuation_range2(days, page, cell_Alpha, cell_num+10, ticker)

get_all(30, 4, "A", 1, "TSLA")

#RSI = AvgGain(평균상승폭)/AvgLoss(평균하락폭)
def rsi(ticker, window):
    data = yf.download(ticker, period="3mo", interval="1d")
    data["Change"] = data["Close"].diff()
    data["Gain"] = data["Change"].clip(lower=0)
    data["Loss"] = -data["Change"].clip(upper=0)
    data["AvgGain"] = data["Gain"].ewm(alpha=1/window, min_periods=window).mean()
    data["AvgLoss"] = data["Loss"].ewm(alpha=1/window, min_periods=window).mean()

    data["RS"] = data["AvgGain"] / data["AvgLoss"]
    data["RSI"] = 100 - (100 / (1 + data["RS"]))

    latest_rsi = data["RSI"].iloc[-1]
    print(f"{ticker} RSI: {latest_rsi:.2f}")

    print(data[["Close", "RSI"]].tail(10))

rsi("TSLA", 14) 
#days_loader(30, 4, "A", 1)
#get_chart(30, 4, "A", 3, "TSLA", "Open")
#get_chart(30, 4, "A", 4, "TSLA", "Close")
#get_chart(30, 4, "A", 6, "TSLA", "High")
#get_chart(30, 4, "A", 7, "TSLA", "Low")
#get_chart(30, 4, "A", 9, "TSLA", "Volume")




