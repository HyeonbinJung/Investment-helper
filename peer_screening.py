import xlwings as xw
import yfinance as yf
import time
import pandas as pd
from sklearn.linear_model import LinearRegression
import numpy as np
from openpyxl.utils import get_column_letter, column_index_from_string

wb = xw.Book(r"C:\Users\sjung\OneDrive\바탕 화면\새 폴더 (4)\trading_helper.xlsx")

def get_marketcap(page, ticker):
    ws = wb.sheets[page]
    stock_target = yf.Ticker(ticker)
    print(stock_target.info.get("marketCap"))
    data = yf.download(ticker, period="3mo", interval="1d")
    print(data)
    data2 = stock_target.quarterly_financials
    print(data2)


#get_marketcap(4, "TSLA")



def insert_excel(page, cell, value):
    ws = wb.sheets[page]
    ws[cell].value = value
    
def get_quarterly_revenue(ticker, target_quarter):
    stock = yf.Ticker(ticker)
    df = stock.quarterly_financials
    if "Total Revenue" not in df.index:
        return str("None");
    target = pd.to_datetime(target_quarter)
    if target in df.columns:
        return (df.loc["Total Revenue", target]);
    else:
         return str("None");


def insert_quarterly_revenue(ticker, cell_Alpha, cell_num):
    col_num = column_index_from_string(cell_Alpha)
    revenue_1q2024 = get_quarterly_revenue(ticker, "2024-03-31")
    revenue_2q2024 = get_quarterly_revenue(ticker, "2024-06-30")
    revenue_3q2024 = get_quarterly_revenue(ticker, "2024-09-30")
    revenue_4q2024 = get_quarterly_revenue(ticker, "2024-12-31")
    revenue_1q2025 = get_quarterly_revenue(ticker, "2025-03-31")
    revenue_2q2025 = get_quarterly_revenue(ticker, "2025-06-30")
    revenue_3q2025 = get_quarterly_revenue(ticker, "2025-09-30")
    for i in range(7):
        alpha = get_column_letter(col_num + i)

        insert_cell = alpha + str(cell_num)
        print (insert_cell)
        if i == 0:
            insert_excel(2, str(insert_cell), revenue_1q2024)
        if i == 1:
            insert_excel(2, str(insert_cell), revenue_2q2024)
        if i == 2:
            insert_excel(2, str(insert_cell), revenue_3q2024)
        if i == 3:
            insert_excel(2, str(insert_cell), revenue_4q2024)
        if i == 4:
            insert_excel(2, str(insert_cell), revenue_1q2025)
        if i == 5:
            insert_excel(2, str(insert_cell), revenue_2q2025)
        if i == 6:
            insert_excel(2, str(insert_cell), revenue_3q2025)

 


    
            
        



